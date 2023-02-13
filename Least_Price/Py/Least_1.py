from openpyxl import load_workbook, Workbook
from Scraping.Least_Price.Form.Form import *
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")

# excel_path = r"D:\Durai\Scraping\Accessories\Save Data's\Final Files\Accessories Price List " + date + ".xlsx"
# excel_path = r"D:\Durai\Scraping\Kitchen_appliances\Save Data\Final Files\Kitchen Appliance Price List " + date + ".xlsx"
# excel_path = r"D:\Durai\Scraping\Mobile\Save Data\Final Files\Mobiles_Price_List " + date + ".xlsx"
# excel_path = r"D:\Durai\Scraping\Laptop\Save Data's\Final Files\Laptop Price Lists " + date + ".xlsx"
# excel_path = r"D:\Durai\Scraping\Tv\Save Data\Final Files\Tv Price List " + date + ".xlsx"
excel_path = r"D:\Durai\Scraping\Tablets\Save Data\Final Files\Tablets Price Lists " + date + ".xlsx"
#
# save_path = r"D:\Durai\Scraping\Least_Price\Save Data\Least Price Accessories Price List " + date + ".xlsx"
# save_path = r"D:\Durai\Scraping\Least_Price\Save Data\Least Price Kitchen Appliance " + date + ".xlsx"
# save_path = r"D:\Durai\Scraping\Least_Price\Save Data\Least Price Mobiles_Price_List " + date + ".xlsx"
# save_path = r"D:\Durai\Scraping\Least_Price\Save Data\Least Price Laptop Price Lists " + date + ".xlsx"
# save_path = r"D:\Durai\Scraping\Least_Price\Save Data\Least Price Tv Price List " + date + ".xlsx"
save_path = r"D:\Durai\Scraping\Least_Price\Save Data\Least Price Tablets " + date + ".xlsx"

wb = load_workbook(excel_path)
ws = wb.active

save_wb = Workbook()
save_ws = save_wb.active

save_ws.cell(row=1, column=1).value = "Product id"
save_ws.cell(row=1, column=2).value = "Item_code"
save_ws.cell(row=1, column=3).value = "Model Name"
save_ws.cell(row=1, column=4).value = "Least Price Brand"
save_ws.cell(row=1, column=5).value = "Least Price"
save_ws.cell(row=1, column=6).value = "Poorvika price"
save_ws.cell(row=1, column=7).value = "Flipkart Price"
save_ws.cell(row=1, column=8).value = "Amazon Price"
save_ws.cell(row=1, column=9).value = "Croma price"
save_ws.cell(row=1, column=10).value = "vijay price"
# save_ws.cell(row=1, column=10).value = "Reliance price"
save_ws.cell(row=1, column=11).value = "Reliance price"

flip = 0
ama = 0
cro = 0
vj = 0
relia = 0
poor = 0
less = 0

for r in range(2, ws.max_row + 1):
# for r in range(2, 100):
    print(r)
    Flipkart = 0
    Amazon = 0
    Croma = 0
    Vijay = 0
    Reliance = 0
    min_value = None
    cell_name = ""

    p_id = ws.cell(row=r, column=1).value
    item_code = ws.cell(row=r, column=2).value
    name = ws.cell(row=r, column=3).value
    p_price = int(ws.cell(row=r, column=4).value)
    f_price = ws.cell(row=r, column=5).value
    a_price = ws.cell(row=r, column=6).value
    c_price = ws.cell(row=r, column=7).value
    v_price = ws.cell(row=r, column=8).value
    # r_price = ws.cell(row=r, column=8).value
    r_price = ws.cell(row=r, column=9).value

    save_ws.cell(row=r, column=1).value = p_id
    save_ws.cell(row=r, column=2).value = item_code
    save_ws.cell(row=r, column=3).value = name

    save_ws.cell(row=r, column=6).value = p_price
    save_ws.cell(row=r, column=7).value = f_price
    save_ws.cell(row=r, column=8).value = a_price
    save_ws.cell(row=r, column=9).value = c_price
    save_ws.cell(row=r, column=10).value = v_price
    # save_ws.cell(row=r, column=10).value = r_price
    save_ws.cell(row=r, column=11).value = r_price

#####################################################################################################################

    if f_price == "NA" and a_price == "NA" and c_price == "NA" and v_price == "NA" and r_price == "NA":
    # if f_price == "NA" and a_price == "NA" and c_price == "NA" and r_price == "NA":
        # Poorvika
        save_ws.cell(row=r, column=4).value = "Poorvika"
        save_ws.cell(row=r, column=5).value = p_price
        save_wb.save(save_path)

    else:
#############################################################################################################

        "Flipkart"
        if f_price != "NA" and p_price >= f_price:
            Flipkart = f_price
        else:
            Flipkart = p_price + 1000

#############################################################################################################

        "Amazon"
        if a_price != "NA" and int(p_price) >= int(a_price):
            Amazon = a_price
        else:
            Amazon = p_price + 1000

#############################################################################################################

        "Croma"
        if c_price != "NA" and p_price >= c_price:
            Croma = c_price
        else:
            Croma = p_price + 1000

#############################################################################################################

        "Vijay"
        if v_price != "NA" and p_price >= v_price:
            Vijay = v_price
        else:
            Vijay = p_price + 1000

#############################################################################################################

        "Reliance"
        if r_price != "NA" and p_price >= r_price:
            Reliance = r_price
        else:
            Reliance = p_price + 1000

#############################################################################################################

        # value = min(Flipkart, Amazon, Croma, Reliance)
        value = min(Flipkart, Amazon, Croma, Vijay, Reliance)

        if value == f_price:
            cell_name = "Flipkart"
            min_value = f_price
            flip = flip + 1

        if value == a_price:
            cell_name = "Amazon"
            min_value = a_price
            ama = ama+1

        if value == c_price:
            cell_name = "Croma"
            min_value = c_price
            cro = cro + 1

        if value == v_price:
            cell_name = "Vijay"
            min_value = v_price
            vj = vj+1

        if value == r_price:
            cell_name = "Reliance"
            min_value = r_price
            relia = relia+1

        if value >= p_price:
            cell_name = "Poorvika"
            min_value = p_price
            poor = poor+1

        if value < p_price and value + (value * 5) / 100 >= p_price:
            cell_name = "poorvika Greater then 5%"
            min_value = value
            less = less+1


        save_ws.cell(row=r, column=4).value = cell_name
        save_ws.cell(row=r, column=5).value = min_value

save_ws1 = "Color"
save_wb.create_sheet(save_ws1)

save_wb[save_ws1]["B2"] = "Brands"
save_wb[save_ws1]["B3"] = "Flipkart"
save_wb[save_ws1]["B4"] = "Amazon"
save_wb[save_ws1]["B5"] = "Croma"
save_wb[save_ws1]["B6"] = "Vijay sale"
save_wb[save_ws1]["B7"] = "Reliance"
save_wb[save_ws1]["B8"] = "Poorvika"
save_wb[save_ws1]["B9"] = "poorvika Greater then 5%"

save_wb[save_ws1]["c2"] = "Totals"
save_wb[save_ws1]["c3"] = flip
save_wb[save_ws1]["c4"] = ama
save_wb[save_ws1]["c5"] = cro
save_wb[save_ws1]["c6"] = vj
save_wb[save_ws1]["c7"] = relia
save_wb[save_ws1]["c8"] = poor
save_wb[save_ws1]["c9"] = less

save_wb.save(save_path)