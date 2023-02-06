from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
import os
import pandas as pd
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")

# 'Product id', 'Item code', 'Model Name', 'Poorvika price','Flipkart Price', 'Amazon Price', 'Croma price', 'Vijay price', 'Reliance price'

wb = Workbook()
ws = wb.active

excel_path = {

    "Accessories": r"D:\Durai\Scraping\Accessories\Save Data's\Final Files\Accessories Price List " + date + ".xlsx",
    "Laptop": r"D:\Durai\Scraping\Laptop\Save Data's\Final Files\Laptop Price Lists " + date + ".xlsx",
    "Mobiles": r"D:\Durai\Scraping\Mobile\Save Data\Final Files\Mobiles_Price_List " + date + ".xlsx",
    "Tablets": r"D:\Durai\Scraping\Tablets\Save Data\Final Files\Tablets Price Lists " + date + ".xlsx",
    "Tv": r"D:\Durai\Scraping\Tv\Save Data\Final Files\Tv Price List " + date + ".xlsx",
    "Kitchen Appliance": r"D:\Durai\Scraping\Kitchen_appliances\Save Data\Final Files\Kitchen Appliance Price List " + date + ".xlsx"

}

thick_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                      bottom=Side(style='thin'))
light_blue = PatternFill(start_color="BAF3F1", end_color="BAF3F1", fill_type="solid")
color = ["74FB65", "FFFF00", "FE3B5B"]
headers = ["Brand", "Poorvika (<)", "Poorvika (=)", "Poorvika (>)", "NA", "Total"]


class Color_Code:
    #############################################################################################
    def Heading(self, **kwargs):
        ws.merge_cells("e6:j6")
        ws["e6"].alignment = Alignment(horizontal="center", vertical="center")
        ws["e6"].font = Font(bold=True)
        ws["e6"] = "Price Comparison - " + kwargs.get("path") + " - " + date
        ws["e6"].fill = PatternFill(start_color="F9AF57", end_color="F9AF57", fill_type="solid")
        ws["e6"].border = thick_border

    #############################################################################################
    def Title(self):
        r = 0
        for head in headers:
            ws.cell(row=7, column=5 + r).value = head
            ws.cell(row=7, column=5 + r).border = thick_border
            ws.cell(row=7, column=5 + r).fill = light_blue
            r = r + 1
            #############################################################################################

    def summer(self):
        cell = ["e", 'f', 'g', 'h']

        ws['e13'] = "Summer"
        ws["f13"] = Total_greater
        ws["g13"] = Total_equal
        ws["h13"] = Total_lesser

        for r in range(0, 4):
            ws[cell[r] + str(13)].fill = light_blue
            ws[cell[r] + str(13)].border = thick_border

    #############################################################################################
    def data_value(self):
        ws.cell(row=l + 8, column=5).value = index[:-6]
        ws.cell(row=l + 8, column=6).value = greater
        ws.cell(row=l + 8, column=7).value = equal
        ws.cell(row=l + 8, column=8).value = lesser
        ws.cell(row=l + 8, column=9).value = data[index].isnull().sum()
        ws.cell(row=l + 8, column=10).value = greater + equal + lesser + data[index].isnull().sum()

        for r in range(5, 10+1):
            ws.cell(row=l + 8, column=r).border = thick_border

        ws.cell(row=l + 8, column=6).fill = PatternFill(start_color="74FB65", end_color="74FB65", fill_type="solid")
        ws.cell(row=l + 8, column=7).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws.cell(row=l + 8, column=8).fill = PatternFill(start_color="FE3B5B", end_color="FE3B5B", fill_type="solid")

    #############################################################################################
    def save_file(self):
        wb.save(r"D:\Durai\Scraping\Least_Price\total_save\Price Comparison " + path + " " + date + ".xlsx")


for path in excel_path:
    print(path)

    data = pd.read_excel(excel_path[path])
    Cc = Color_Code()
    Cc.Heading(path=path)
    Cc.Title()

    l = 0
    Total_greater = 0
    Total_equal = 0
    Total_lesser = 0
    for index in data.columns:
        if index not in ["Product id", "Item code", "Model Name", "Poorvika price"]:
            # print(index)
            equal = 0
            greater = 0
            lesser = 0

            for k, r in data.iterrows():
                if r.iloc[3] == r.iloc[4 + l]:
                    equal = equal + 1
                elif r.iloc[3] < r.iloc[4 + l]:
                    greater = greater + 1
                elif r.iloc[3] > r.iloc[4 + l]:
                    lesser = lesser + 1
                Cc.data_value()

            Total_greater += greater
            Total_equal += equal
            Total_lesser += lesser
            l = l + 1
            Cc.summer()

        Cc.save_file()
# #################################################################################################
