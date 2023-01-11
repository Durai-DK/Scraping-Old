from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.styles.borders import Border,Side
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")

#######################################################################################################################

excel_path = r"D:\Durai\Scraping\Home_appliances\Save Date's\Final Files\Home_appliances " + date + ".xlsx"

Name = "Home appliances"

#######################################################################################################################

wb = load_workbook(excel_path)
ws = wb.active

dk_wb = Workbook()
dk_ws = dk_wb.active

#######################################################################################################################

dk_ws.merge_cells("d6:h6")
dk_ws["d6"] = "Price Comparison - " + Name + " - " + date
dk_ws["d6"].alignment = Alignment(horizontal="center",vertical="center")
dk_ws["d6"].font = Font(bold=True)

dk_ws["d7"] = "Brand"
dk_ws["e7"] = "Poorvika  <  "
dk_ws["f7"] = "Poorvika  =  "
dk_ws["g7"] = "Poorvika  >  "
dk_ws["h7"] = "NA"

dk_ws["d8"] = "Sathiya"
dk_ws["d9"] = "Vasanth"
dk_ws["d10"] = "Darling"
dk_ws["d11"] = "Vivek's "
dk_ws["d12"] = "Croma"
dk_ws["d13"] = "Amazon"
dk_ws["d14"] = "Flipkart"
dk_ws["d15"] = "Reliance"
dk_ws["d16"] = "Summary"

Flipkart_1 = 0
Flipkart_2 = 0
Flipkart_3 = 0
Flipkart_4 = 0

Amazon_1 = 0
Amazon_2 = 0
Amazon_3 = 0
Amazon_4 = 0

Croma_1 = 0
Croma_2 = 0
Croma_3 = 0
Croma_4 = 0

Vivek_1 = 0
Vivek_2 = 0
Vivek_3 = 0
Vivek_4 = 0

Reliance_1 = 0
Reliance_2 = 0
Reliance_3 = 0
Reliance_4 = 0

Sathiya_1 = 0
Sathiya_2 = 0
Sathiya_3 = 0
Sathiya_4 = 0


Vasanth_1 = 0
Vasanth_2 = 0
Vasanth_3 = 0
Vasanth_4 = 0

Darling_1 = 0
Darling_2 = 0
Darling_3 = 0
Darling_4 = 0

for r in range(2, ws.max_row + 1):

    p_price = int(ws.cell(row=r, column=4).value)
    s_price = ws.cell(row=r, column=5).value
    va_price = ws.cell(row=r, column=6).value
    d_price = ws.cell(row=r, column=7).value
    vi_price = ws.cell(row=r, column=8).value
    c_price = ws.cell(row=r, column=9).value
    a_price = ws.cell(row=r, column=10).value
    f_price = ws.cell(row=r, column=11).value
    r_price = ws.cell(row=r, column=12).value

#####################################################################################################################

    if p_price != "NA" and f_price != "NA" and p_price < f_price:
        Flipkart_1 = Flipkart_1 + 1
    elif p_price != "NA" and f_price != "NA" and p_price == f_price:
        Flipkart_2 = Flipkart_2 + 1
    elif p_price != "NA" and f_price != "NA" and p_price > f_price:
        Flipkart_3 = Flipkart_3 + 1
    else:
        Flipkart_4 = Flipkart_4 + 1

  ########################################################################################

    if p_price != "NA" and a_price != "NA" and p_price < a_price:
        Amazon_1 = Amazon_1 + 1
    elif p_price != "NA" and a_price != "NA" and p_price == a_price:
        Amazon_2 = Amazon_2 + 1
    elif p_price != "NA" and a_price != "NA" and p_price > a_price:
        Amazon_3 = Amazon_3 + 1
    else:
        Amazon_4 = Amazon_4 + 1

  ########################################################################################

    if p_price != "NA" and c_price != "NA" and p_price < c_price:
        Croma_1 = Croma_1 + 1
    elif p_price != "NA" and c_price != "NA" and p_price == c_price:
        Croma_2 = Croma_2 + 1
    elif p_price != "NA" and c_price != "NA" and p_price > c_price:
        Croma_3 = Croma_3 + 1
    else:
        Croma_4 = Croma_4 + 1

  ########################################################################################

    if p_price != "NA" and vi_price != "NA" and p_price < vi_price:
        Vivek_1 = Vivek_1 + 1
    elif p_price != "NA" and vi_price != "NA" and p_price == vi_price:
        Vivek_2 = Vivek_2 + 1
    elif p_price != "NA" and vi_price != "NA" and p_price > vi_price:
        Vivek_3 = Vivek_3 + 1
    else:
        Vivek_4 = Vivek_4 + 1

  ########################################################################################

    if p_price != "NA" and r_price != "NA" and p_price < r_price:
        Reliance_1 = Reliance_1 + 1
    elif p_price != "NA" and r_price != "NA" and p_price == r_price:
        Reliance_2 = Reliance_2 + 1
    elif p_price != "NA" and r_price != "NA" and p_price > r_price:
        Reliance_3 = Reliance_3 + 1
    else:
        Reliance_4 = Reliance_4 + 1

  ########################################################################################

    if p_price != "NA" and s_price != "NA" and p_price < s_price:
        Sathiya_1 = Sathiya_1 + 1
    elif p_price != "NA" and s_price != "NA" and p_price == s_price:
        Sathiya_2 = Sathiya_2 + 1
    elif p_price != "NA" and s_price != "NA" and p_price > s_price:
        Sathiya_3 = Sathiya_3 + 1
    else:
        Sathiya_4 = Sathiya_4 + 1

  ########################################################################################

    if p_price != "NA" and va_price != "NA" and p_price < va_price:
        Vasanth_1 = Vasanth_1 + 1
    elif p_price != "NA" and va_price != "NA" and p_price == va_price:
        Vasanth_2 = Vasanth_2 + 1
    elif p_price != "NA" and va_price != "NA" and p_price > va_price:
        Vasanth_3 = Vasanth_3 + 1
    else:
        Vasanth_4 = Vasanth_4 + 1

  ########################################################################################

    if p_price != "NA" and d_price != "NA" and p_price < d_price:
        Darling_1 = Darling_1 + 1
    elif p_price != "NA" and d_price != "NA" and p_price == d_price:
        Darling_2 = Darling_2 + 1
    elif p_price != "NA" and d_price != "NA" and p_price > d_price:
        Darling_3 = Darling_3 + 1
    else:
        Darling_4 = Darling_4 + 1
#####################################################################################################################

print()
dk_ws["e8"] = Sathiya_1
dk_ws["f8"] = Sathiya_2
dk_ws["g8"] = Sathiya_3
dk_ws["h8"] = Sathiya_4
print("Sathya Done")

dk_ws["e9"] = Vasanth_1
dk_ws["f9"] = Vasanth_2
dk_ws["g9"] = Vasanth_3
dk_ws["h9"] = Vasanth_4
print("Vasanth Done")

dk_ws["e10"] = Darling_1
dk_ws["f10"] = Darling_1
dk_ws["g10"] = Darling_1
dk_ws["h10"] = Darling_1
print("Darling Done")

dk_ws["e11"] = Vivek_1
dk_ws["f11"] = Vivek_2
dk_ws["g11"] = Vivek_3
dk_ws["h11"] = Vivek_4
print("Vivek Done")

dk_ws["e12"] = Croma_1
dk_ws["f12"] = Croma_2
dk_ws["g12"] = Croma_3
dk_ws["h12"] = Croma_4
print("Croma Done")

dk_ws["e13"] = Amazon_1
dk_ws["f13"] = Amazon_2
dk_ws["g13"] = Amazon_3
dk_ws["h13"] = Amazon_4
print("Amazon Done")

dk_ws["e14"] = Flipkart_1
dk_ws["f14"] = Flipkart_2
dk_ws["g14"] = Flipkart_3
dk_ws["h14"] = Flipkart_4
print("Flipkart Done")

dk_ws["e15"] = Reliance_1
dk_ws["f15"] = Reliance_2
dk_ws["g15"] = Reliance_3
dk_ws["h15"] = Reliance_4
print("Reliance Done")

#####################################################################################################################

dk_ws["e16"] = int(dk_ws["e8"].value) + int(dk_ws["e9"].value) + int(dk_ws["e10"].value) + int(dk_ws["e11"].value) + int(dk_ws["e12"].value) + int(dk_ws["e13"].value) + int(dk_ws["e14"].value) + int(dk_ws["e15"].value)

dk_ws["f16"] = int(dk_ws["f8"].value) + int(dk_ws["f9"].value) + int(dk_ws["f10"].value) + int(dk_ws["f11"].value) + int(dk_ws["f12"].value) + int(dk_ws["f13"].value) + int(dk_ws["f14"].value) + int(dk_ws["f15"].value)

dk_ws["g16"] = int(dk_ws["g8"].value) + int(dk_ws["g9"].value) + int(dk_ws["g10"].value) + int(dk_ws["g11"].value) + int(dk_ws["g12"].value) + int(dk_ws["g13"].value) + int(dk_ws["g14"].value) + int(dk_ws["g15"].value)

dk_ws["h16"] = int(dk_ws["h8"].value) + int(dk_ws["h9"].value) + int(dk_ws["h10"].value) + int(dk_ws["h11"].value) + int(dk_ws["h12"].value) + int(dk_ws["h13"].value) + int(dk_ws["h14"].value) + int(dk_ws["h15"].value)

# Green = "74FB65"
# Yellow = "FFFF00"
# red = "FE3B5B"
# Blue = "BAF3F1"
# Brown = "F9AF57"

color = ["74FB65","FFFF00","FE3B5B"]

for c in range(5,8):
    for r in range(8, 16):
        cell_header = dk_ws.cell(r, c)
        cell_header.fill = PatternFill(start_color=color[c-5], end_color=color[c-5], fill_type="solid")

for c1 in range(4,9):
    cell1 = dk_ws.cell(7,c1)
    cell1.fill = PatternFill(start_color="BAF3F1", end_color="BAF3F1", fill_type="solid")

    cell2 = dk_ws.cell(16,c1)
    cell2.fill = PatternFill(start_color="BAF3F1", end_color="BAF3F1", fill_type="solid")

dk_ws["d6"].fill = PatternFill(start_color="F9AF57", end_color="F9AF57", fill_type="solid")


for cell in dk_ws._cells.values():
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    cell.border = thin_border

dk_wb.save(r"D:\Durai\Scraping\Least_Price\total_save\Price Comparison " + Name + " " + date + ".xlsx")
