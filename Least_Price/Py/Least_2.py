from openpyxl import load_workbook, Workbook
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")


######################################################################################################################

excel_path = r"D:\Durai\Scraping\Home_appliances\Save Date's\Final Files\Home_appliances " + date + ".xlsx"

######################################################################################################################

save_path = r"D:\Durai\Scraping\Least_Price\Save Data\Least Home appliances " + date + ".xlsx"

######################################################################################################################

wb = load_workbook(excel_path)
ws = wb.active

save_wb = Workbook()
save_ws = save_wb.active

save_ws.cell(row=1, column=1).value = "Product id"
save_ws.cell(row=1, column=2).value = "Item_code"
save_ws.cell(row=1, column=3).value = "Model Name"
save_ws.cell(row=1, column=4).value = "Least Price Brand"
save_ws.cell(row=1, column=5).value = "Least Price"
save_ws.cell(row=1, column=6).value = "Poorvika price"
save_ws.cell(row=1, column=7).value = "Sathiya Price"
save_ws.cell(row=1, column=8).value = "Vasanth & CO Price"
save_ws.cell(row=1, column=9).value = "Darling price"
save_ws.cell(row=1, column=10).value = "Vivek's price"
save_ws.cell(row=1, column=11).value = "Croma price"
save_ws.cell(row=1, column=12).value = "Amazon price"
save_ws.cell(row=1, column=13).value = "Flipkart price"
save_ws.cell(row=1, column=14).value = "Reliance price"

flip = 0
ama = 0
cro = 0
va = 0
vi = 0
da = 0
poor = 0
less = 0
sa = 0
rel = 0


for r in range(2, ws.max_row + 1):
    # for r in range(2, 10):
    print(r)
    Sathiya_value = 0
    vasanth_value = 0
    Darling_value = 0
    vivek_value = 0
    croma_value = 0
    Amazon_value = 0
    Flipkart_value = 0
    Reliance_value = 0
    min_value = None
    cell_name = ""

    p_id = ws.cell(row=r, column=1).value
    item_code = ws.cell(row=r, column=2).value
    name = ws.cell(row=r, column=3).value
    p_price = int(ws.cell(row=r, column=4).value)
    s_price = ws.cell(row=r, column=5).value
    va_price = ws.cell(row=r, column=6).value
    d_price = ws.cell(row=r, column=7).value
    vi_price = ws.cell(row=r, column=8).value
    c_price = ws.cell(row=r, column=9).value
    a_price = ws.cell(row=r, column=10).value
    f_price = ws.cell(row=r, column=11).value
    r_price = ws.cell(row=r, column=12).value

    save_ws.cell(row=r, column=1).value = p_id
    save_ws.cell(row=r, column=2).value = item_code
    save_ws.cell(row=r, column=3).value = name

    save_ws.cell(row=r, column=6).value = p_price
    save_ws.cell(row=r, column=7).value = s_price
    save_ws.cell(row=r, column=8).value = va_price
    save_ws.cell(row=r, column=9).value = d_price
    save_ws.cell(row=r, column=10).value = vi_price
    save_ws.cell(row=r, column=11).value = c_price
    save_ws.cell(row=r, column=12).value = a_price
    save_ws.cell(row=r, column=13).value = f_price
    save_ws.cell(row=r, column=14).value = r_price

    if s_price == "NA" and va_price == "NA" and d_price == "NA" and vi_price == "NA" and c_price == "NA" and a_price == "NA" and f_price == "NA" and r_price == "NA":
        # Poorvika
        save_ws.cell(row=r, column=4).value = "Poorvika"
        save_ws.cell(row=r, column=5).value = p_price

        save_wb.save(save_path)
    else:

        "Sathya"
        if s_price != "NA" and p_price >= s_price:
            Sathiya_value = s_price
        else:
            Sathiya_value = p_price + 1000

#####################################################################################################################

        "vasanth"
        if va_price != "NA" and p_price >= va_price:
            vasanth_value = va_price
        else:
            vasanth_value = p_price + 1000

#############################################################################################################

        "Darling"
        if d_price != "NA" and p_price >= d_price:
            Darling_value = d_price
        else:
            Darling_value = p_price + 1000

#############################################################################################################

        "viveks"
        if vi_price != "NA" and p_price >= vi_price:
            vivek_value = vi_price
        else:
            vivek_value = p_price + 1000

#############################################################################################################

        "Croma"
        if c_price != "NA" and p_price >= c_price:
            croma_value = c_price
        else:
            croma_value = p_price + 1000

#############################################################################################################

        "Amazon"
        if a_price != "NA" and p_price >= a_price:
            Amazon_value = a_price
        else:
            Amazon_value = p_price + 1000

#############################################################################################################

        "Flipkart"
        if f_price != "NA" and p_price >= f_price:
            Flipkart_value = f_price
        else:
            Flipkart_value = p_price + 1000

#############################################################################################################

        "Reliance"
        if r_price != "NA" and p_price >= r_price:
            Reliance_value = r_price
        else:
            Reliance_value = p_price + 1000

#############################################################################################################
        value = min(Flipkart_value, Amazon_value, croma_value, vasanth_value, Reliance_value,vivek_value,Sathiya_value,Darling_value, p_price)


        if value == Sathiya_value:
            cell_name = "sathya"
            min_value = s_price
            sa = sa+1

        if value == vasanth_value:
            cell_name = "vasanth"
            min_value = va_price
            va = va +1

        if value == Darling_value:
            cell_name = "Darling"
            min_value = d_price
            da = da+1

        if value == vivek_value:
            cell_name = "viveks"
            min_value = vi_price
            vi = vi+1

        if value == croma_value:
            cell_name = "Croma"
            min_value = c_price
            cro = cro+1

        if value == Amazon_value:
            cell_name = "Amazon"
            min_value = a_price
            ama = ama+1

        if value == Flipkart_value:
            cell_name = "Flipkart"
            min_value = f_price
            flip = flip+1

        if value == Reliance_value:
            cell_name = "Reliance"
            min_value = r_price
            rel = rel + 1

        if value >= p_price:
            cell_name = "Poorvika"
            min_value = p_price
            poor = poor+1

        if value < p_price and value + (value * 5) / 100 >= p_price:
            cell_name = "poorvika Greater then 5%"
            min_value = value
            less = less+1

        save_ws.cell(row=r, column=4).value = cell_name
        save_ws.cell(row=r, column=5).value = min_value

save_ws1 = "Color"
save_wb.create_sheet(save_ws1)

save_wb[save_ws1]["B2"] = "Brands"
save_wb[save_ws1]["B3"] = "Flipkart"
save_wb[save_ws1]["B4"] = "Amazon"
save_wb[save_ws1]["B5"] = "Croma"
save_wb[save_ws1]["B6"] = "Vasanth"
save_wb[save_ws1]["B7"] = "Reliance"
save_wb[save_ws1]["B8"] = "sathya"
save_wb[save_ws1]["B9"] = "Viveks"
save_wb[save_ws1]["B10"] = "Darling"
save_wb[save_ws1]["B11"] = "Poorvika"
save_wb[save_ws1]["B12"] = "poorvika Greater then 5%"

save_wb[save_ws1]["c2"] = "Totals"
save_wb[save_ws1]["c3"] = flip
save_wb[save_ws1]["c4"] = ama
save_wb[save_ws1]["c5"] = cro
save_wb[save_ws1]["c6"] = va
save_wb[save_ws1]["c7"] = rel
save_wb[save_ws1]["c8"] = sa
save_wb[save_ws1]["c9"] = vi
save_wb[save_ws1]["c10"] = da
save_wb[save_ws1]["c11"] = poor
save_wb[save_ws1]["c12"] = less

save_wb.save(save_path)
