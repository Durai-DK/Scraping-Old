from openpyxl import load_workbook,Workbook
from Scraping.Least_Price.Form.Datas import path
from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.styles.borders import Border,Side
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")

dk_wb = Workbook()
dk_ws = dk_wb.active


class forms:
    def __init__(self, ws=None):
        self.ws = ws

    def woerkbook(self, **kwargs):
        if kwargs.get('head') == "Accessories":
            wb = load_workbook(path["Accessories"]['Data'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Laptop":
            wb = load_workbook(path["Laptop"]['Data'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Mobile":
            wb = load_workbook(path["Mobile"]['Data'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Tv":
            wb = load_workbook(path["Tv"]['Data'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Tablets":
            wb = load_workbook(path["Tv"]['Data'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "KA":
            wb = load_workbook(path["KA"]['Data'] + date + ".xlsx")
            ws = wb.active

    def run_data(self):
        dk_ws.merge_cells("d6:h6")
        dk_ws["d6"] = "Price Comparison - " + self.Name + " - " + date
        dk_ws["d6"].alignment = Alignment(horizontal="center",vertical="center")
        dk_ws["d6"].font = Font(bold=True)

        dk_ws["d7"] = "Brand"
        dk_ws["e7"] = "Poorvika <"
        dk_ws["f7"] = "Poorvika ="
        dk_ws["g7"] = "Poorvika >"
        dk_ws["h7"] = "NA"

        dk_ws["d8"] = "Flipkart"
        dk_ws["d9"] = "Amazon"
        dk_ws["d10"] = "Croma"
        dk_ws["d11"] = "Vijay"
        dk_ws["d12"] = "Reliance"
        dk_ws["d13"] = "Summary"


        Flipkart_1 = 0
        Flipkart_2 = 0
        Flipkart_3 = 0
        Flipkart_4 = 0

        Amazon_1 = 0
        Amazon_2 = 0
        Amazon_3 = 0
        Amazon_4 = 0

        Croma_1 = 0
        Croma_2 = 0
        Croma_3 = 0
        Croma_4 = 0

        Vijay_1 = 0
        Vijay_2 = 0
        Vijay_3 = 0
        Vijay_4 = 0

        Reliance_1 = 0
        Reliance_2 = 0
        Reliance_3 = 0
        Reliance_4 = 0

        for r in range(2, self.ws.max_row + 1):
            p_price = int(self.ws.cell(row=r, column=4).value)
            f_price = self.ws.cell(row=r, column=5).value
            a_price = self.ws.cell(row=r, column=6).value
            c_price = self.ws.cell(row=r, column=7).value
            v_price = self.ws.cell(row=r, column=8).value
            r_price = self.ws.cell(row=r, column=9).value


            if p_price != "NA" and f_price != "NA" and p_price < f_price:
                Flipkart_1 = Flipkart_1 + 1
            elif p_price != "NA" and f_price != "NA" and p_price == f_price:
                Flipkart_2 = Flipkart_2 + 1
            elif p_price != "NA" and f_price != "NA" and p_price > f_price:
                Flipkart_3 = Flipkart_3 + 1
            else:
                Flipkart_4 = Flipkart_4 + 1


            if p_price != "NA" and a_price != "NA" and p_price < a_price:
                Amazon_1 = Amazon_1 + 1
            elif p_price != "NA" and a_price != "NA" and p_price == a_price:
                Amazon_2 = Amazon_2 + 1
            elif p_price != "NA" and a_price != "NA" and p_price > a_price:
                Amazon_3 = Amazon_3 + 1
            else:
                Amazon_4 = Amazon_4 + 1


            if p_price != "NA" and c_price != "NA" and p_price < c_price:
                Croma_1 = Croma_1 + 1
            elif p_price != "NA" and c_price != "NA" and p_price == c_price:
                Croma_2 = Croma_2 + 1
            elif p_price != "NA" and c_price != "NA" and p_price > c_price:
                Croma_3 = Croma_3 + 1
            else:
                Croma_4 = Croma_4 + 1


            if p_price != "NA" and v_price != "NA" and p_price < v_price:
                Vijay_1 = Vijay_1 + 1
            elif p_price != "NA" and v_price != "NA" and p_price == v_price:
                Vijay_2 = Vijay_2 + 1
            elif p_price != "NA" and v_price != "NA" and p_price > v_price:
                Vijay_3 = Vijay_3 + 1
            else:
                Vijay_4 = Vijay_4 + 1


            if p_price != "NA" and r_price != "NA" and p_price < r_price:
                Reliance_1 = Reliance_1 + 1
            elif p_price != "NA" and r_price != "NA" and p_price == r_price:
                Reliance_2 = Reliance_2 + 1
            elif p_price != "NA" and r_price != "NA" and p_price > r_price:
                Reliance_3 = Reliance_3 + 1
            else:
                Reliance_4 = Reliance_4 + 1

        print(Flipkart_1)
        dk_ws["e8"] = Flipkart_1
        print(Flipkart_2)
        dk_ws["f8"] = Flipkart_2
        print(Flipkart_3)
        dk_ws["g8"] = Flipkart_3
        print(Flipkart_4)
        dk_ws["h8"] = Flipkart_4

        print()
        print(Amazon_1)
        dk_ws["e9"] = Amazon_1
        print(Amazon_2)
        dk_ws["f9"] = Amazon_2
        print(Amazon_3)
        dk_ws["g9"] = Amazon_3
        print(Amazon_4)
        dk_ws["h9"] = Amazon_4

        print()
        print(Croma_1)
        dk_ws["e10"] = Croma_1
        print(Croma_2)
        dk_ws["f10"] = Croma_2
        print(Croma_3)
        dk_ws["g10"] = Croma_3
        print(Croma_4)
        dk_ws["h10"] = Croma_4

        print()
        print(Vijay_1)
        dk_ws["e11"] = Vijay_1
        print(Vijay_2)
        dk_ws["f11"] = Vijay_2
        print(Vijay_3)
        dk_ws["g11"] = Vijay_3
        print(Vijay_4)
        dk_ws["h11"] = Vijay_4

        print()
        print(Reliance_1)
        dk_ws["e12"] = Reliance_1
        print(Reliance_2)
        dk_ws["f12"] = Reliance_2
        print(Reliance_3)
        dk_ws["g12"] = Reliance_3
        print(Reliance_4)
        dk_ws["h12"] = Reliance_4


        dk_ws["e13"] = int(dk_ws["e8"].value) + int(dk_ws["e9"].value) + int(dk_ws["e10"].value) + int(
            dk_ws["e11"].value) + int(dk_ws["e12"].value)

        dk_ws["f13"] = int(dk_ws["f8"].value) + int(dk_ws["f9"].value) + int(dk_ws["f10"].value) + int(
            dk_ws["f11"].value) + int(dk_ws["f12"].value)

        dk_ws["g13"] = int(dk_ws["g8"].value) + int(dk_ws["g9"].value) + int(dk_ws["g10"].value) + int(
            dk_ws["g11"].value) + int(dk_ws["g12"].value)

        dk_ws["h13"] = int(dk_ws["h8"].value) + int(dk_ws["h9"].value) + int(dk_ws["h10"].value) + int(
            dk_ws["h11"].value) + int(dk_ws["h12"].value)

        color = ["74FB65", "FFFF00", "FE3B5B"]

        for c in range(5, 8):
            for r in range(8, 13):
                cell_header = dk_ws.cell(r, c)
                cell_header.fill = PatternFill(start_color=color[c - 5], end_color=color[c - 5], fill_type="solid")

        for c1 in range(4, 9):
            cell1 = dk_ws.cell(7, c1)
            cell1.fill = PatternFill(start_color="BAF3F1", end_color="BAF3F1", fill_type="solid")

            cell2 = dk_ws.cell(13, c1)
            cell2.fill = PatternFill(start_color="BAF3F1", end_color="BAF3F1", fill_type="solid")

        dk_ws["d6"].fill = PatternFill(start_color="F9AF57", end_color="F9AF57", fill_type="solid")

        for cell in dk_ws._cells.values():
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.border = thin_border

        dk_wb.save(r"D:\Durai\Scraping\Least_Price\total_save\Price Comparison " + self.Name + " " + date + ".xlsx")

#####################################################################################################################