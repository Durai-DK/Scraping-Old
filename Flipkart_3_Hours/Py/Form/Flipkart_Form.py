from selenium import webdriver
from openpyxl import load_workbook,Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.service import Service
import time
import datetime
Date = datetime.datetime.now().strftime("%d-%m-%Y")

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")

old_wb = load_workbook(r"D:\Durai\Scraping\Flipkart_3_Hours\Urls\Flipkart Scraping.xlsx")
old_ws = old_wb.active

wb = Workbook()
ws = wb.active

##############################################################################################################

class Model:

    def heading(self):

        ws.cell(row=1, column=1).value = "Item Code"
        ws.cell(row=1, column=2).value = "Name"
        ws.cell(row=1, column=3).value = "Flipkart Urls"
        ws.cell(row=1, column=4).value = "Flipkart Name"
        ws.cell(row=1, column=5).value = "Flipkart Price"

    def excel(self,**kwargs):

        ws.cell(row=kwargs.get('range'), column=1).value = old_ws.cell(row=kwargs.get('range'), column=1).value
        ws.cell(row=kwargs.get('range'), column=2).value = old_ws.cell(row=kwargs.get('range'), column=2).value
        ws.cell(row=kwargs.get('range'), column=3).value = old_ws.cell(row=kwargs.get('range'), column=3).value

    def initial(self,**kwargs):

        try:

            f_name = driver.find_element(By.CLASS_NAME, "B_NuCI").text
            print(f_name)
            ws.cell(row=kwargs.get('range'), column=4).value = f_name

            f_price = driver.find_element(By.CLASS_NAME, "_30jeq3").text
            print(f_price[1:])
            ws.cell(row=kwargs.get('range'), column=5).value = f_price[1:]

        except:
            pass

    def seller(self,**kwargs):

        try:
            d = 7
            driver.find_element(By.CLASS_NAME, "_1_xoMS").click()
            time.sleep(2)

            for r1 in driver.find_elements(By.CLASS_NAME, "_2Y3EWJ"):
                s_name = driver.find_element(By.CLASS_NAME, "_3enH42").text
                ws.cell(row=kwargs.get('range'), column=d).value = s_name
                d = d + 1

                s_price = driver.find_element(By.CLASS_NAME, "_30jeq3").text
                ws.cell(row=kwargs.get('range'), column=d).value = s_price
                d = d + 1

        except:
            pass

    def save(self,**kwargs):

        wb.save(r"D:\Durai\Scraping\Flipkart_3_Hours\Save Data\Scraping Sheets\flipkart_Scraping " + str(kwargs.get('path')) + " " + Date + ".xlsx")

    def flipkart(self,**kwargs):

        Model.heading(self)

        for r in range(kwargs.get("Start"),kwargs.get("End")+1):
            print("")
            print(r)

            Model.excel(self,range=r)

            driver.get(url=old_ws.cell(row=r, column=3).value)
            time.sleep(2)

            Model.initial(self,range=r)

            Model.seller(self,range=r)
            Model.save(self,range=r,path=kwargs.get('Path'))

        driver.quit()