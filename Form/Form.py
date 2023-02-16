from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import datetime
from selenium import webdriver
from openpyxl import Workbook
import time
import pymongo
#
# date = datetime.date.today().strftime("%d-%m-%Y")
#
# s = Service(r"D:\Durai\Driver\chromedriver.exe")
# driver = webdriver.Chrome(service=s)
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

date = datetime.datetime.now().strftime("%d-%m-%Y")
s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)

wb = Workbook()
ws = wb.active
ws.title = "Accessories"

my_host = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = my_host["Poorvikawebsite"]
mycol = mydb[date]

########################################################################################################################
class Model:

    def __init__(self, url, row_num, heading):

        self.url = url
        self.row_num = row_num
        self.heading = heading

########################################################################################################################
    def product_page(self):

        for box in driver.find_elements(By.CLASS_NAME, "product-card_card__description__2LoI_"):
            name = box.find_element(By.TAG_NAME, "b").text
            price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
            print(name)
            print(price[2:])

            ws.cell(row=self.row_num, column=1).value = name
            ws.cell(row=self.row_num, column=2).value = price[2:]
            ws.cell(row=self.row_num, column=3).value = self.heading

            mydict = {
                "name": name,
                "price": price[2:],
                "cat": self.heading
            }

            x = mycol.insert_one(mydict)

            self.product_save(head=self.heading)
            self.row_num = self.row_num + 1

########################################################################################################################
    def product_page_list(self, **kwargs):
        print("Title :", kwargs.get('head'))
        page = ""

        if kwargs.get('head') == "Mobile & Accessories":
            page = 32

        elif kwargs.get('head') == "Computer & Laptop Accessories":
            page = 7

        elif kwargs.get('head') == "Tab & Ipad Accessories":
            page = 3

        elif kwargs.get('head') == "TV & Audio Accessories":
            page = 10

        elif kwargs.get('head') == "Smart Technology":
            page = 5

        elif kwargs.get('head') == "Laptops":
            page = 5

        elif kwargs.get('head') == "Tablets":
            page = 5

        elif kwargs.get('head') == "Mobiles":
            page = 17

        elif kwargs.get('head') == "Tv":
            page = 5

        elif kwargs.get('head') == "Kitchen Appliances":
            page = 20

        self.product(page=page)

########################################################################################################################

    def product(self, **kwargs):
        print(kwargs.get('page'))
        # try:
        for r in range(1, int(kwargs.get('page'))):
            print("                     ")
            print("Page range:" + str(r))
            driver.get(self.url + str(r))
            time.sleep(5)
            self.product_page()
        # except:
        #     pass

        print(self.heading + " Complete")
        print("#" * 60)

########################################################################################################################
    def product_1(self):
        for r in range(1,2):
            print(r)

            driver.get(self.url)

            for box in driver.find_elements(By.CLASS_NAME, "product-card_card__description__2LoI_"):
                name = box.find_element(By.TAG_NAME, "b").text
                price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
                print(name)
                print(price[2:])

                ############### excel ##################################
                ws.cell(row=self.row_num, column=1).value = name
                ws.cell(row=self.row_num, column=2).value = price[2:]
                ws.cell(row=self.row_num, column=3).value = self.heading

                ############## mogodb ##################################
                mydict = {
                    "name": name,
                    "price":price[2:],
                    "cat":self.heading
                }
                x = mycol.insert_one(mydict)

                self.row_num = self.row_num + 1
                self.product_save(head=self.heading)
            print(self.heading + " Complete")
            print("#" * 60)

########################################################################################################################

    def product_num(self):
        return self.row_num

########################################################################################################################
    def product_save(self,**kwargs):

        if kwargs.get('head') == "Laptops":
            wb.save(r"D:\Durai\Scraping\Laptop\Save Data's\Poorvika Files\Laptop " + date + ".xlsx")

        elif kwargs.get('head') == "Tablets":
            wb.save(r"D:\Durai\Scraping\Tablets\Save Data\Poorvilka Files\Tablets " + date + ".xlsx")

        elif kwargs.get('head') == "Tv":
            wb.save(r"D:\Durai\Scraping\Tv\Save Data\Poorvilka Files\Tv " + date + ".xlsx")

        elif kwargs.get('head') == "Mobiles":
            wb.save(r"D:\Durai\Scraping\Mobile\Save Data\Poorvilka Files\Mobiles " + date + ".xlsx")

        elif kwargs.get('head') == "Kitchen Appliances":
            wb.save(r"D:\Durai\Scraping\Kitchen_appliances\Save Data\Poorvika Files\kitchen Appliances " + date + ".xlsx")

        else:
            wb.save(r"D:\Durai\Scraping\Accessories\Save Data's\Poorvika Files\Accessories " + date + ".xlsx")

########################################################################################################################
