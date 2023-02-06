from selenium.webdriver.common.by import By
import time
import datetime
from selenium import webdriver
from openpyxl import load_workbook, Workbook
from Scraping.Form.Product_Details import product, competitor, path
from selenium.common.exceptions import NoSuchElementException
date = datetime.date.today().strftime("%#d-%#m-%Y")
print(date)

driver = webdriver.Chrome(executable_path=path["Driver"]["Chrome"])


new_wb = Workbook()
new_ws = new_wb.active


class PriceCompression:
    def __init__(self,  row_num=None, ws=None):
        self.row_num = row_num
        self.ws = ws

    def heading(self):
        new_ws.cell(row=1, column=1).value = product[3] + " " + product[0]
        new_ws.cell(row=1, column=2).value = competitor[0] + " " + product[1]
        new_ws.cell(row=1, column=3).value = competitor[1] + " " + product[0]
        new_ws.cell(row=1, column=4).value = competitor[1] + " " + product[1]
        new_ws.cell(row=1, column=5).value = competitor[1] + " " + product[2]
        new_ws.cell(row=1, column=6).value = competitor[2] + " " + product[0]
        new_ws.cell(row=1, column=7).value = competitor[2] + " " + product[1]
        new_ws.cell(row=1, column=8).value = competitor[2] + " " + product[2]
        new_ws.cell(row=1, column=9).value = competitor[3] + " " + product[0]
        new_ws.cell(row=1, column=10).value = competitor[3] + " " + product[1]
        new_ws.cell(row=1, column=11).value = competitor[3] + " " + product[2]
        new_ws.cell(row=1, column=12).value = competitor[4] + " " + product[0]
        new_ws.cell(row=1, column=13).value = competitor[4] + " " + product[1]
        new_ws.cell(row=1, column=14).value = competitor[4] + " " + product[2]
        new_ws.cell(row=1, column=15).value = competitor[5] + " " + product[0]
        new_ws.cell(row=1, column=16).value = competitor[5] + " " + product[1]
        new_ws.cell(row=1, column=17).value = competitor[5] + " " + product[2]

    def flipkart(self):
        if self.ws.cell(row=self.row_num, column=3).value != "N/A":
            print("#" * 100)
            print("Web Site :Flipkart")
            print("Web Link :", self.ws.cell(row=self.row_num, column=3).value)
            new_ws.cell(row=self.row_num, column=5).value = self.ws.cell(row=self.row_num, column=3).value

            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=3).value)
                time.sleep(1)

                new_ws.cell(row=self.row_num, column=3).value = driver.find_element(By.TAG_NAME, "h1").text
                new_ws.cell(row=self.row_num, column=4).value = driver.find_element(By.CLASS_NAME, "_30jeq3").text[1:]
                print("Product Name :", driver.find_element(By.TAG_NAME, "h1").text)
                print("Product Price :", driver.find_element(By.CLASS_NAME, "_30jeq3").text[1:])
            except:
                pass

    def amazon(self):
        if self.ws.cell(row=self.row_num, column=4).value != "N/A":
            print("#" * 100)
            print("Web Site :Amazon")
            print("Web Link :", self.ws.cell(row=self.row_num, column=4).value)
            new_ws.cell(row=self.row_num, column=8).value = self.ws.cell(row=self.row_num, column=4).value
            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=4).value)
                driver.implicitly_wait(3)

                name = driver.find_element(By.ID, "productTitle").text
                print("Product Name :", name)
                new_ws.cell(row=self.row_num, column=6).value = name

                try:

                    price = driver.find_element(By.ID, "apex_desktop")
                    price1 = price.find_element(By.CLASS_NAME, "a-price-whole")
                    print("Amazon Price 1 :", price1.text)
                    new_ws.cell(row=self.row_num, column=7).value = price1.text

                except:
                    pass

                try:

                    pric = driver.find_element(By.ID, "apex_desktop")
                    pric2 = pric.find_element(By.CLASS_NAME, "apexPriceToPay")
                    print("Amazon Price 2 :", pric2.text)
                    new_ws.cell(row=self.row_num, column=7).value = pric2.text

                except:
                    pass

            except:
                pass

    def croma(self):
        if self.ws.cell(row=self.row_num, column=5).value != "N/A":
            print("#" * 100)
            print("Web Site :Croma")
            print("Web Link :", self.ws.cell(row=self.row_num, column=5).value)
            new_ws.cell(row=self.row_num, column=11).value = self.ws.cell(row=self.row_num, column=5).value
            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=5).value)
                time.sleep(3)

                print("Product Name :", driver.find_element(By.TAG_NAME, "h1").text)
                new_ws.cell(row=self.row_num, column=9).value = driver.find_element(By.TAG_NAME, "h1").text

                # for tit in driver.find_elements(By.CLASS_NAME, "outer-product-pricebox"):
                #     for price in tit.find_elements(By.CLASS_NAME, "pdp-cp-price"):
                #         for price1 in price.find_elements(By.CLASS_NAME, 'new-price'):
                price2 = driver.find_element(By.ID, 'pdp-product-price')
                print("Product Price :", price2.text[1:])
                new_ws.cell(row=self.row_num, column=10).value = price2.text[1:]

            except:
                pass

    def vijay_sale(self):
        if self.ws.cell(row=self.row_num, column=6).value != "N/A":
            print("#" * 100)
            print("Web Site :vijay sale")
            print("Web Link :", self.ws.cell(row=self.row_num, column=6).value)
            new_ws.cell(row=self.row_num, column=14).value = self.ws.cell(row=self.row_num, column=6).value
            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=6).value)
                driver.implicitly_wait(3)
                print("Product Name :", driver.find_element(By.TAG_NAME, "h1").text)
                new_ws.cell(row=self.row_num, column=12).value = driver.find_element(By.TAG_NAME, "h1").text
                try:
                    if driver.find_element(By.ID, "ContentPlaceHolder1_fillprice").text != None:
                        # print("Nothing")
                        try:
                            print("Web Site : Vijay Sale")
                            print("Vijay Sale Price 1 :", driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text)
                            price = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text
                            new_ws.cell(row=self.row_num, column=13).value = price

                        except NoSuchElementException:
                            print("Web Site : Vijay Sale")
                            print("Vijay Sale Price 2 :", driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text)
                            price = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text
                            new_ws.cell(row=self.row_num, column=13).value = price

                except:
                    pass
            except:
                pass

    def reliance(self):
        if self.ws.cell(row=self.row_num, column=7).value != "N/A":
            print("#" * 100)
            print("Web Site :Reliance")
            print("Web Link :", self.ws.cell(row=self.row_num, column=7).value)
            new_ws.cell(row=self.row_num, column=17).value = self.ws.cell(row=self.row_num, column=7).value
            try:
                driver.get(url=self.ws.cell(row=self.row_num, column=7).value)
                time.sleep(1)

                try:
                    print("Product Name :", driver.find_element(By.TAG_NAME, 'h1').text)
                    new_ws.cell(row=self.row_num, column=15).value = driver.find_element(By.TAG_NAME, 'h1').text
                    for pri in driver.find_elements(By.CLASS_NAME, "pdp__priceSection"):
                        price = pri.find_element(By.CLASS_NAME, 'pdp__priceSection__priceListText')
                        print("Reliance Price 1 = ", price.text)
                        new_ws.cell(row=self.row_num, column=16).value = price.text
                except:
                    pass

                try:
                    new_ws.cell(row=self.row_num, column=15).value = driver.find_element(By.TAG_NAME, 'h1').text
                    for pric in driver.find_elements("blk__sm__6"):
                        price1 = pric.find_element(By.CLASS_NAME, 'pdp__offerPrice')
                        print("Reliance Price 2 = ", price1.text)
                        new_ws.cell(row=self.row_num, column=16).value = price1.text
                except:
                    pass

            except:
                pass


class RunCompression:

    def run__all(self, **kwargs):
        if kwargs.get('head') == "Accessories":
            wb = load_workbook(path["Accessories"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Laptop":
            wb = load_workbook(path["Laptop"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Mobile":
            wb = load_workbook(path["Mobile"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Tv":
            wb = load_workbook(path["Tv"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "Tablets":
            wb = load_workbook(path["Tv"]['Url'] + date + ".xlsx")
            ws = wb.active

        elif kwargs.get('head') == "KA":
            wb = load_workbook(path["KA"]['Url'] + date + ".xlsx")
            ws = wb.active

        ph = PriceCompression()
        ph.heading()

        for r in range(kwargs.get('start'), kwargs.get('end')+1):
            print("")
            print(r)
            new_ws.cell(row=r, column=1).value = ws.cell(row=r, column=1).value
            new_ws.cell(row=r, column=2).value = ws.cell(row=r, column=2).value
            pc = PriceCompression(row_num=r, ws=ws)
            pc.flipkart()
            pc.amazon()
            pc.croma()
            pc.vijay_sale()
            pc.reliance()
            self.save(head=kwargs.get('head'), path=kwargs.get('path'))
        driver.quit()

    def save(self, **kwargs):
        if kwargs.get("head") == "Accessories":
            path_save = path["Accessories"]['Save']
            path_num = str(kwargs.get('path'))
            new_wb.save(path_save + path_num + " Price List " + date + ".xlsx")

        elif kwargs.get("head") == "Laptop":
            new_wb.save(path["Laptop"]['Save'] + date + ".xlsx")

        elif kwargs.get("head") == "Mobile":
            path_save = path["Mobile"]['Save']
            path_num = str(kwargs.get('path'))
            new_wb.save(path_save + path_num + " Price List " + date + ".xlsx")

        elif kwargs.get("head") == "Tv":
            new_wb.save(path["Tv"]['Save'] + date + ".xlsx")

        elif kwargs.get("head") == "Tablets":
            new_wb.save(path["Tablets"]['Save'] + date + ".xlsx")

        if kwargs.get("head") == "KA":
            path_save = path["KA"]['Save']
            path_num = str(kwargs.get('path'))
            new_wb.save(path_save + path_num + " Price List " + date + ".xlsx")
