from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import datetime
from selenium import webdriver
from openpyxl import Workbook
import time

date = datetime.date.today().strftime("%d-%m-%Y")

s = Service(r"D:\Durai\Driver\chromedriver.exe")
driver = webdriver.Chrome(service=s)

wb = Workbook()
ws = wb.active
ws.title = "Tv"

def Tv_Scraping():
    l = 1
    for r in range(1, 4):
        print("                  ")
        print("Page = " + str(r))


        time.sleep(2)
        driver.get("https://www.poorvika.com/s?catagories=categories.lvl1%3A%3D%5B%60TV+%26+Audio+%3E+Television%60%5D&stock_status=stock_status%3A%3D%5B%60In+Stock%60%5D&page=" + str(r))

        for box in driver.find_elements(By.CLASS_NAME, "product-card_card__description__2LoI_"):
            name = box.find_element(By.TAG_NAME, "b").text
            print(name)
            ws.cell(row=l, column=1).value = name

            price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
            print(price)
            ws.cell(row=l, column=2).value = price[2:]

            ws.cell(row=l, column=3).value = "Tv"

            wb.save("D:\Durai\Scraping\Tv\Save Data\Poorvilka Files\Tv " + date + ".xlsx")
            l = l + 1
