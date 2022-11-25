
from openpyxl import load_workbook, Workbook
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")

# excel_path = r"D:\Durai\Scraping\Accessories\Save Data's\Final Files\Accessories Price List " + date + ".xlsx"
# excel_path = r"D:\Durai\Scraping\Laptop\Save Data's\Final Files\Laptop Price Lists " + date + ".xlsx"
excel_path = r"D:\Durai\Scraping\Mobile\Save Data\Final Files\Mobiles_Price_List " + date + ".xlsx"
# excel_path = r"D:\Durai\Scraping\Tv\Save Data\Final Files\Tv Price List " + date + ".xlsx"
# excel_path = r"D:\Durai\Scraping\Kitchen_appliances\Save Data\Kitchen Appliance " + date + ".xlsx"

# excel_path = r"D:\Durai\Scraping\Home_appliances\Save Date's\Final Files\Home_appliances " + date + ".xlsx"


######################################################################################################################

# save_path = r"D:\Durai\Scraping\Accessories\Save Data's\Least Price\Least Price Accessories Price List " + date + ".xlsx"
# save_path = r"D:\Durai\Scraping\Accessories\Save Data's\Least Price\Least Price Laptop Price Lists " + date + ".xlsx"
save_path = r"D:\Durai\Scraping\Accessories\Save Data's\Least Price\Least Price Mobiles_Price_List " + date + ".xlsx"
# save_path = r"D:\Durai\Scraping\Accessories\Save Data's\Least Price\Least Price Tv Price List " + date + ".xlsx"
# save_path = r"D:\Durai\Scraping\Accessories\Save Data's\Least Price\Least Price Kitchen Appliance " + date + ".xlsx"

# save_path = r"D:\Durai\Scraping\Home_appliances\Save Date's\Least Price\Home_appliances " + date + ".xlsx"

wb = load_workbook(excel_path)
ws = wb.active

save_wb = Workbook()
save_ws = save_wb.active

save_ws.cell(row=1, column=1).value = "Product id"
save_ws.cell(row=1, column=2).value = "Item_code"
save_ws.cell(row=1, column=3).value = "Model Name"
save_ws.cell(row=1, column=4).value = "Least Price Brand"
save_ws.cell(row=1, column=5).value = "Least Price"
save_ws.cell(row=1, column=6).value = "Poorvika price"
save_ws.cell(row=1, column=7).value = "Flipkart Price"
save_ws.cell(row=1, column=8).value = "Amazon Price"
save_ws.cell(row=1, column=9).value = "Croma price"
save_ws.cell(row=1, column=10).value = "vijay price"
save_ws.cell(row=1, column=11).value = "Reliance price"

for r in range(2, ws.max_row + 1):
# for r in range(2, 10):
    print(r)
    Flipkart_value = None
    Amazon_value = None
    croma_value = None
    Vijay_value = None
    Reliance_value = None
    cell_name = ""

    p_id = ws.cell(row=r, column=1).value
    item_code = ws.cell(row=r, column=2).value
    name = ws.cell(row=r, column=3).value
    p_price = int(ws.cell(row=r, column=4).value)
    f_price = ws.cell(row=r, column=5).value
    a_price = ws.cell(row=r, column=6).value
    c_price = ws.cell(row=r, column=7).value
    v_price = ws.cell(row=r, column=8).value
    r_price = ws.cell(row=r, column=9).value

    save_ws.cell(row=r, column=1).value = p_id
    save_ws.cell(row=r, column=2).value = item_code
    save_ws.cell(row=r, column=3).value = name

    save_ws.cell(row=r, column=6).value = p_price
    save_ws.cell(row=r, column=7).value = f_price
    save_ws.cell(row=r, column=8).value = a_price
    save_ws.cell(row=r, column=9).value = c_price
    save_ws.cell(row=r, column=10).value = v_price
    save_ws.cell(row=r, column=11).value = r_price

    if f_price == "NA" and a_price == "NA" and c_price == "NA" and v_price == "NA" and r_price == "NA":
        # Poorvika
        save_ws.cell(row=r, column=4).value = "Poorvika"
        save_ws.cell(row=r, column=5).value = p_price

        save_wb.save(save_path)
    else:

        "Flipkart"
        if f_price != "NA" and p_price >= f_price:
            Flipkart_value = f_price
        else:
            Flipkart_value = p_price + 1000
            
        "Amazon"
        if a_price != "NA" and p_price >= a_price:
            Amazon_value = a_price
        else:
            Amazon_value = p_price + 1000

        "Croma"
        if c_price != "NA" and p_price >= c_price:
            croma_value = c_price
        else:
            croma_value = p_price + 1000

        "Vijay"
        if v_price != "NA" and p_price >= v_price:
            Vijay_value = v_price
        else:
            Vijay_value = p_price + 1000
            
        "Reliance"
        if r_price != "NA" and p_price >= r_price:
            Reliance_value = r_price
        else:
            Reliance_value = p_price + 1000

        min_value = min(Flipkart_value, Amazon_value, croma_value, Vijay_value, Reliance_value, p_price)

        if min_value == p_price:
            cell_name += "Poorvika "
        if min_value == Flipkart_value:
            cell_name += "Flipkart "
        if min_value == Amazon_value:
            cell_name += "Amazon "
        if min_value == croma_value:
            cell_name += "Croma "
        if min_value == Vijay_value:
            cell_name += "Vijay "
        if min_value == Reliance_value:
            cell_name += "Reliance "

        save_ws.cell(row=r, column=4).value = cell_name
        save_ws.cell(row=r, column=5).value = min_value

        save_wb.save(save_path)



