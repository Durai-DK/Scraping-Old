from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
import datetime
from selenium import webdriver
from openpyxl import Workbook
import time


date = datetime.date.today().strftime("%d-%m-%Y")

s = ChromeService(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)


wb = Workbook()
ws = wb.active
ws.title = "Mobile"

l = 1
for r in range(1, 5):
    print(" ")
    print(r)

    driver.get(url = "https://www.poorvika.com/s?categories=categories.lvl1%3A%3D%5B%60Mobiles+%26+Accessories+%3E+Mobiles%60%5D&stock_status=stock_status%3A%3D%5B%60In+Stock%60%5D&page=" + str(r))
    time.sleep(2)

    for box in driver.find_elements(By.CLASS_NAME, "product-card_card__description__2LoI_"):
        name = box.find_element(By.TAG_NAME, "b").text
        price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
        print(name)
        print(price[2:])

        ws.cell(row=l, column=1).value = name
        ws.cell(row=l, column=2).value = price[2:]

        l = l + 1


    wb.save(r"D:\Durai\Scraping\Mobile\Mobile test " + date + ".xlsx")