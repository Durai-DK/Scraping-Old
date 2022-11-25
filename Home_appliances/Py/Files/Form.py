from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import datetime
hp_wb = load_workbook(r"D:\Durai\Scraping\Home_appliances\Web Url\Urls.xlsx")
hp_ws = hp_wb.active
print("No Of Rows",hp_ws.max_row)

date = datetime.date.today().strftime("%#d-%#m-%Y")

options = webdriver.ChromeOptions()

prefs = {'profile.default_content_setting_values': {'notifications': 2, 'auto_select_certificate': 2, 'fullscreen': 2,
                                                    'mouselock': 2, 'mixed_script': 2, 'media_stream': 2,
                                                    'media_stream_mic': 2, 'media_stream_camera': 2,
                                                    'protocol_handlers': 2,
                                                    'ppapi_broker': 2, 'automatic_downloads': 2, 'midi_sysex': 2,
                                                    'push_messaging': 2, 'ssl_cert_decisions': 2,
                                                    'metro_switch_to_desktop': 2,
                                                    'protected_media_identifier': 2, 'app_banner': 2,
                                                    'site_engagement': 2,'popups': 2,
                                                    'durable_storage': 2}}

s = Service(r"D:/Durai/Driver/chromedriver.exe")
driver = webdriver.Chrome(service=s, options=options)

save_wb = Workbook()
save_ws = save_wb.active

save_ws.cell(row=1, column=1).value = "Product id"
save_ws.cell(row=1, column=2).value = "Item Code"
save_ws.cell(row=1, column=3).value = "Model"
save_ws.cell(row=1, column=4).value = "Poorvika"
save_ws.cell(row=1, column=5).value = "Sathiya"
save_ws.cell(row=1, column=6).value = "Vasanth & CO"
save_ws.cell(row=1, column=7).value = "Darling retail"
save_ws.cell(row=1, column=8).value = "Vivek's"
save_ws.cell(row=1, column=9).value = "Croma"
save_ws.cell(row=1, column=10).value = "Amazon"
save_ws.cell(row=1, column=11).value = "Flipkart"
save_ws.cell(row=1, column=12).value = "Reliance"
save_ws.cell(row=1, column=13).value = "Tata Cliq"


class Model:
    def __init__(self,row_num=None, ws=None):
        self.row_num = row_num
        self.ws = ws

    def Sathya(self):
        if hp_ws.cell(row=self.row_num, column=4).value != "N/A":
            print(self.row_num)
            print(hp_ws.cell(row=self.row_num, column=4).value)
            try:
                driver.get(url=hp_ws.cell(row=self.row_num, column=4).value)
                try:
                    price = driver.find_element(By.CLASS_NAME,"pd-price-block").text
                    print("Sathiya = ", price[2:])
                    save_ws.cell(row=self.row_num, column=5).value = price[2:]
                except:

                    for box in driver.find_elements(By.CLASS_NAME, "price-box"):
                        price1 = box.find_element(By.CLASS_NAME,"product-price").text
                        print("Sathiya = ", price1)
                        save_ws.cell(row=self.row_num, column=2).value = price1
            except:
                pass

    def Vasanth(self):
        if hp_ws.cell(row=self.row_num, column=5).value != "N/A":
            # print("Vasanth & Co")
            print(hp_ws.cell(row=self.row_num, column=5).value)
            try:
                driver.get(url=hp_ws.cell(row=self.row_num, column=5).value)
                vas = driver.find_element(By.CLASS_NAME, "price-new").text
                print("Vasanth & Co = ", vas[2:])
                save_ws.cell(row=self.row_num, column=6).value = vas[2:]
            except:
                pass

    def Darling(self):
        if hp_ws.cell(row=self.row_num, column=6).value != 'N/A':
            # print("Darling")
            print(hp_ws.cell(row=self.row_num, column=6).value)
            try:
                driver.get(url=hp_ws.cell(row=self.row_num, column=6).value)
                for title in driver.find_elements(By.ID, "detail-price"):
                    price = title.find_element(By.CLASS_NAME, "price-sale").text
                    print("Darling = ", price[4:])
                    save_ws.cell(row=self.row_num, column=7).value = price[4:]
            except:
                pass

    def viveks(self):
        if hp_ws.cell(row=self.row_num, column=7).value != "N/A":
            # print("Vivek's")
            print(hp_ws.cell(row=self.row_num, column=7).value)
            try:
                driver.get(url=hp_ws.cell(row=self.row_num, column=7).value)
                for price in driver.find_elements(By.CLASS_NAME, "product-info-price"):
                    price1 = driver.find_element(By.CLASS_NAME, "price").text
                    print(price1)
                    print("Vivek's = ", price1[1:])
                    save_ws.cell(row=self.row_num, column=8).value = price1[1:]
            except:
                pass

    def Croma(self):
        if hp_ws.cell(row=self.row_num, column=8).value != "N/A":
            # print("Croma")
            print(hp_ws.cell(row=self.row_num, column=8).value)
            try:
                driver.get(url=hp_ws.cell(row=self.row_num, column=8).value)
                for price in driver.find_elements(By.CLASS_NAME, "outer-product-pricebox"):
                    for pric in price.find_elements(By.CLASS_NAME, "main-product-price"):
                        price1 = pric.find_element(By.CLASS_NAME, "new-price").text
                        print("Croma = ", price1[1:])
                        save_ws.cell(row=self.row_num, column=9).value = price1[1:]
            except:
                pass

    def Amazon(self):
        if hp_ws.cell(row=self.row_num, column=9).value != "N/A":
            # print("Amazon")
            print(hp_ws.cell(row=self.row_num, column=9).value)
            try:
                try:
                    driver.get(url=hp_ws.cell(row=self.row_num, column=9).value)
                    p1 = driver.find_element(By.ID, "apex_desktop")
                    price = p1.find_element(By.CLASS_NAME, "apexPriceToPay").text
                    print("Amazon = ", price)
                    save_ws.cell(row=self.row_num, column=10).value = price

                except:
                    price3 = driver.find_element(By.ID, "apex_desktop")
                    price4 = price3.find_element(By.CLASS_NAME, "a-price-whole")
                    print("Amazon = ", price4.text)
                    save_ws.cell(row=self.row_num, column=10).value = price4.text

            except:
                pass

    def flipkart(self):
        if hp_ws.cell(row=self.row_num, column=10).value != "N/A":
            # print("Flipkart")
            print(hp_ws.cell(row=self.row_num, column=10).value)
            try:
                driver.get(url=hp_ws.cell(row=self.row_num, column=10).value)
                for price in driver.find_elements(By.CLASS_NAME, "_1AtVbE"):
                    flip = price.find_element(By.CLASS_NAME, "_16Jk6d").text
                    print("Flipkart = ", flip[1:])
                    save_ws.cell(row=self.row_num, column=11).value = flip[1:]
            except:
                pass

    def Reliance(self):
        if hp_ws.cell(row=self.row_num, column=11).value != "N/A":
            # print("Reliance")
            print(driver.get(url=hp_ws.cell(row=self.row_num, column=11).value))
            try:
                driver.get(url=hp_ws.cell(row=self.row_num, column=11).value)
                price = driver.find_element(By.CLASS_NAME, "pdp__offerPrice").text
                print("Reliance = ", price[1:])
                save_ws.cell(row=self.row_num, column=12).value = price[1:]
            except:
                pass

    def Tata(self):
        if hp_ws.cell(row=self.row_num, column=12).value != "N/A":
            # print("Tata Cliq")
            print(hp_ws.cell(row=self.row_num, column=12).value)
        try:
            driver.get(url=hp_ws.cell(row=self.row_num, column=12).value)
            price = driver.find_element(By.CLASS_NAME, "PriceSection__discounted-price-block").text
            print("Tata = ", price[1:])
            save_ws.cell(row=self.row_num, column=13).value = price[1:]
        except:
            pass

    def save(self, **kwargs):

        path = r"D:\Durai\Scraping\Home_appliances\Save Date's\Save Files\files\Home Application "
        path_num = str(kwargs.get('path'))
        save_wb.save(path + path_num + " " + date + ".xlsx")


class RunHome:

    def run(self, **kwargs):
        for r in range(kwargs.get('start'), kwargs.get('end')):
            # print(r)
            save_ws.cell(row=r, column=1).value = hp_ws.cell(row=r, column=1).value
            save_ws.cell(row=r, column=2).value = hp_ws.cell(row=r, column=2).value
            save_ws.cell(row=r, column=3).value = hp_ws.cell(row=r, column=3).value

            pc = Model(row_num=r)
            pc.Sathya()
            pc.Vasanth()
            pc.Darling()
            pc.viveks()
            pc.Croma()
            pc.Amazon()
            pc.flipkart()
            pc.Reliance()
            pc.Tata()
            pc.save(path=kwargs.get('path'))


