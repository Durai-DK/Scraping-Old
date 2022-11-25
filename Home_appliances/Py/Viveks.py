from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
import datetime
ac_wb = load_workbook(r"D:\Durai\Scraping\Home_appliances\Web Url\Urls.xlsx")
ac_ws = ac_wb.active
print("No Of Rows",ac_ws.max_row)
print("No Of Columns",ac_ws.max_column)


date = datetime.date.today().strftime("%#d-%#m-%Y")

options = webdriver.ChromeOptions()

prefs = {'profile.default_content_setting_values': {'cookies': 2, 'images': 2, 'popups': 2, 'geolocation':2,
# prefs = {'profile.default_content_setting_values': {'notifications': 2, 'auto_select_certificate': 2, 'fullscreen': 2,'popups': 2,
                                                    'mouselock': 2, 'mixed_script': 2, 'media_stream': 2,'plugins': 2,
                                                    'media_stream_mic': 2, 'media_stream_camera': 2,
                                                    'protocol_handlers': 2,
                                                    'ppapi_broker': 2, 'automatic_downloads': 2, 'midi_sysex': 2,
                                                    'push_messaging': 2, 'ssl_cert_decisions': 2,
                                                    'metro_switch_to_desktop': 2,
                                                    'protected_media_identifier': 2, 'app_banner': 2,
                                                    'site_engagement': 2,
                                                    'durable_storage': 2}}
options.add_experimental_option('prefs', prefs)
# options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")

save_wb = Workbook()
save_ws = save_wb.active

s = Service(r"D:\Durai\Driver\chromedriver.exe")
driver = webdriver.Chrome(service=s, options=options)

save_ws.cell(row=1, column=1).value = "Product Id"
save_ws.cell(row=1, column=2).value = "Item Code"
save_ws.cell(row=1, column=3).value = "Model"
save_ws.cell(row=1, column=4).value = "Poorvika"
save_ws.cell(row=1, column=5).value = "Vivek's Price"
save_ws.cell(row=1, column=6).value = "Vivek's Url"

for r in range(2, ac_ws.max_row+1):
# for r in range(2, 101):

    save_ws.cell(row=r, column=1).value = ac_ws.cell(row=r, column=1).value
    save_ws.cell(row=r, column=2).value = ac_ws.cell(row=r, column=2).value
    save_ws.cell(row=r, column=3).value = ac_ws.cell(row=r, column=3).value

    if ac_ws.cell(row=r, column=7).value != "NA":
        save_ws.cell(row=r, column=6).value = ac_ws.cell(row=r, column=7).value
        print("#" * 150)
        print("Vivek's")
        print('Range: ', r)
        print(ac_ws.cell(row=r, column=7).value)

        try:

            driver.get(url=ac_ws.cell(row=r, column=7).value)
            time.sleep(3)

            for price1 in driver.find_elements(By.CLASS_NAME, "product-info-price"):
                for price2 in price1.find_elements(By.CLASS_NAME, "price-box"):
                    price3 = price2.find_element(By.CLASS_NAME, "special-price")
                    print("Vivek's Price 1 = ", price3.text[1:])
                    save_ws.cell(row=r, column=5).value = price3.text[1:]

        except:
            pass


    save_wb.save(r"D:\Durai\Scraping\Home_appliances\Save Date's\Save Files\Total Scraping\Home Application Viveks " + date + ".xlsx")

driver.quit()

