from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import datetime
from selenium import webdriver
from openpyxl import Workbook
import time

date = datetime.date.today().strftime("%d-%m-%Y")

s = Service(r"D:\Durai\Driver\chromedriver.exe")
driver = webdriver.Chrome(service=s)

wb = Workbook()
ws = wb.active
ws.title = "Accessories"


########################################################################################################################
class Model:

    def __init__(self, url, row_num, heading):

        self.url = url
        self.row_num = row_num
        self.heading = heading

########################################################################################################################
    def product_page(self):

        for box in driver.find_elements(By.CLASS_NAME, "product-card_card__description__2LoI_"):
            name = box.find_element(By.TAG_NAME, "b").text
            price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
            print(name)
            print(price[2:])

            ws.cell(row=self.row_num, column=1).value = name
            ws.cell(row=self.row_num, column=2).value = price[2:]
            ws.cell(row=self.row_num, column=3).value = self.heading
            self.product_save(head=self.heading)
            self.row_num = self.row_num + 1

########################################################################################################################
    def product_page_list(self, **kwargs):
        print("Title :", kwargs.get('head'))
        page = ""
        if kwargs.get('head') == "Mobile & Accessories":
            # page = 3
            page = 30
        elif kwargs.get('head') == "Computer & Laptop Accessories":
            # page = 3
            page = 8
        elif kwargs.get('head') == "Tab & Ipad Accessories":
            # page = 3
            page = 4
        elif kwargs.get('head') == "TV & Audio Accessories":
            # page = 3
            page = 11
        elif kwargs.get('head') == "Smart Technology":
            # page = 3
            page = 7
        elif kwargs.get('head') == "Laptops":
            page = 5
            # page = 2
        self.product(page=page)

########################################################################################################################
    def product(self, **kwargs):
        print(kwargs.get('page'))
        for r in range(1, int(kwargs.get('page'))):
            print("                     ")
            print("Page range:" + str(r))
            time.sleep(2)
            driver.get(self.url + str(r))
            self.product_page()

        print(self.heading + " Complete")
        print("#" * 60)

########################################################################################################################
    def product_1(self):
        for r in range(1,2):
            print(r)

            driver.get(self.url)

            for box in driver.find_elements(By.CLASS_NAME, "product-card_card__description__2LoI_"):
                name = box.find_element(By.TAG_NAME, "b").text
                price = box.find_element(By.CLASS_NAME, "whitespace-nowrap").text
                print(name)
                print(price[2:])

                ws.cell(row=self.row_num, column=1).value = name
                ws.cell(row=self.row_num, column=2).value = price[2:]
                ws.cell(row=self.row_num, column=3).value = self.heading
                self.row_num = self.row_num + 1
                self.product_save(head=self.heading)
            print(self.heading + " Complete")
            print("#" * 60)

########################################################################################################################

    def product_num(self):
        return self.row_num

########################################################################################################################
    def product_save(self,**kwargs):

        if kwargs.get('head') == "Laptops":
            wb.save(r"D:\Durai\Scraping\Laptop\Save Data's\Poorvika Files\Laptop " + date + ".xlsx")
        else:
            wb.save(r"D:\Durai\Scraping\Accessories\Save Data's\Poorvika Files\Accessories " + date + ".xlsx")

########################################################################################################################
