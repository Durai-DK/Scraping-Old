from selenium import webdriver
from openpyxl import load_workbook, Workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import datetime

# options = webdriver.ChromeOptions()

date = datetime.date.today().strftime("%#d-%#m-%Y")

ac_wb = load_workbook(r"D:\Durai\Scraping\Accessories\Save Data's\Scraping Files\Total_Scraping\New folder\Try.xlsx")
ac_ws = ac_wb.active

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")

wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = "Model"
ws.cell(row=1, column=2).value = "Price"
ws.cell(row=1, column=3).value = "Urls"
driver.maximize_window()
options = Options()


for r in range(3, 45):

    ws.cell(row=r, column=1).value = ac_ws.cell(row=r, column=1).value
    # ws.cell(row=r, column=3).value = ac_ws.cell(row=r, column=2).value

    if ac_ws.cell(row=r, column=2).value != "N/A":

        ws.cell(row=r, column=3).value = ac_ws.cell(row=r, column=2).value
        print("#" * 150)
        print('Range : ', r)
        print("Url : ", ac_ws.cell(row=r, column=2).value)

        try:
            driver.get(url=ac_ws.cell(row=r, column=2).value)
            time.sleep(5)

            try:
                for pri in driver.find_elements(By.CLASS_NAME, "pdp__priceSection"):
                    price = pri.find_element(By.CLASS_NAME, 'pdp__priceSection__priceListText')
                    print("Reliance Price 1 = ", price.text)
                    ws.cell(row=r, column=2).value = price.text
            except:
                pass

            try:
                for pric in driver.find_elements("blk__sm__6"):
                    for pric1 in pric.find_elements(By.CLASS_NAME, "pdp__priceSection"):
                        pric2 = pric1.find_element(By.CLASS_NAME, 'pdp__offerPrice')
                        print("Reliance Price 2 = ", pric2.text)
                        ws.cell(row=r, column=2).value = pric2.text
            except:
                pass

        except:
            pass


    wb.save(r"D:\Durai\Scraping\Accessories\Save Data's\Scraping Files\Total_Scraping\New folder\Product .xlsx")

driver.quit()
